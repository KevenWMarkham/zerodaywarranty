"""Scenario library lookup + registration.

Reads the APEX scenario library (``docs/reference/scenario-library.csv``, the
grep-able extract of ``docs/reference/APEX-Scenario-Chains.xlsx``) so scenarios
can be searched from the CLI, and registers this repository's own
``service/*/scenario.yaml`` into the library when it is missing — writing both
the CSV and the Excel workbook.

The CSV is always updated; the ``.xlsx`` is updated when ``openpyxl`` is
installed (it ships in the dev dependency group).
"""

from __future__ import annotations

import contextlib
import csv
from pathlib import Path

from pydantic import BaseModel, ConfigDict

from zero_day_warranty.manifest import load_agent, load_scenario

REPO_ROOT = Path(__file__).resolve().parents[2]
LIBRARY_CSV = REPO_ROOT / "docs" / "reference" / "scenario-library.csv"
LIBRARY_XLSX = REPO_ROOT / "docs" / "reference" / "APEX-Scenario-Chains.xlsx"
SERVICE_DIR = REPO_ROOT / "service"

LIBRARY_SHEET = "Scenario Library"
LIBRARY_HEADER = [
    "#",
    "Scenario ID",
    "Title",
    "Service Code",
    "Domain",
    "Schemas",
    "Brief (the moment)",
    "KPI / Outcome",
    "Featured?",
    "Device(s)",
]


class ScenarioRow(BaseModel):
    """One row of the scenario library."""

    model_config = ConfigDict(extra="forbid")

    index: int | None = None
    scenario_id: str
    title: str = ""
    service_code: str = ""
    domain: str = ""
    schemas: str = ""
    brief: str = ""
    kpi: str = ""
    featured: str = "Catalog"
    devices: str = ""

    @property
    def industry(self) -> str:
        """Industry prefix from the service code (e.g. ``axle``)."""
        return self.service_code.split("-")[0].lower() if self.service_code else ""

    def to_csv_row(self) -> list[str]:
        """Render in ``LIBRARY_HEADER`` order."""
        return [
            "" if self.index is None else str(self.index),
            self.scenario_id,
            self.title,
            self.service_code,
            self.domain,
            self.schemas,
            self.brief,
            self.kpi,
            self.featured,
            self.devices,
        ]


# ---------------------------------------------------------------------------
# Read
# ---------------------------------------------------------------------------


def load_library(path: str | Path = LIBRARY_CSV) -> list[ScenarioRow]:
    """Load the scenario library from the CSV extract."""
    rows: list[ScenarioRow] = []
    with Path(path).open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader, None)
        for raw in reader:
            if not raw or not any(raw):
                continue
            cells = (raw + [""] * len(LIBRARY_HEADER))[: len(LIBRARY_HEADER)]
            idx = cells[0].strip()
            rows.append(
                ScenarioRow(
                    index=int(idx) if idx.isdigit() else None,
                    scenario_id=cells[1].strip(),
                    title=cells[2].strip(),
                    service_code=cells[3].strip(),
                    domain=cells[4].strip(),
                    schemas=cells[5].strip(),
                    brief=cells[6].strip(),
                    kpi=cells[7].strip(),
                    featured=cells[8].strip() or "Catalog",
                    devices=cells[9].strip(),
                )
            )
    _ = header
    return rows


def search(
    rows: list[ScenarioRow],
    *,
    term: str | None = None,
    industry: str | None = None,
    domain: str | None = None,
) -> list[ScenarioRow]:
    """Filter library rows by free-text term, industry prefix, and/or domain."""
    out = rows
    if industry:
        out = [r for r in out if r.industry == industry.lower()]
    if domain:
        out = [r for r in out if domain.lower() in r.domain.lower()]
    if term:
        t = term.lower()
        out = [
            r
            for r in out
            if t in r.scenario_id.lower()
            or t in r.title.lower()
            or t in r.brief.lower()
            or t in r.kpi.lower()
            or t in r.service_code.lower()
        ]
    return out


def find(rows: list[ScenarioRow], scenario_id: str) -> ScenarioRow | None:
    """Return the row with this scenario id, or None."""
    return next((r for r in rows if r.scenario_id == scenario_id), None)


# ---------------------------------------------------------------------------
# This repo's own scenarios
# ---------------------------------------------------------------------------


def repo_scenarios(service_dir: str | Path = SERVICE_DIR) -> list[ScenarioRow]:
    """Build library rows from every ``service/*/scenario.yaml`` in the repo."""
    out: list[ScenarioRow] = []
    for scenario_path in sorted(Path(service_dir).glob("*/scenario.yaml")):
        sc = load_scenario(scenario_path)
        schemas: set[str] = set()
        for ref in sc.agents:
            agent = load_agent(scenario_path.parent / ref.config)
            schemas.update(agent.schemas_read)
            schemas.update(agent.schemas_write)
        out.append(
            ScenarioRow(
                scenario_id=sc.scenario_id,
                title=sc.title,
                service_code=sc.service_code,
                domain=sc.domain,
                schemas=" · ".join(sorted(schemas)),
                brief=" ".join(sc.moment.split())[:240],
                kpi=" ".join(sc.kpi.split())[:240],
                featured="Featured" if sc.featured else "Catalog",
            )
        )
    return out


def missing_from_library(
    library: list[ScenarioRow], candidates: list[ScenarioRow]
) -> list[ScenarioRow]:
    """Candidates whose scenario_id is not already in the library."""
    known = {r.scenario_id for r in library}
    return [c for c in candidates if c.scenario_id not in known]


# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------


def append_to_csv(rows: list[ScenarioRow], path: str | Path = LIBRARY_CSV) -> int:
    """Append rows to the CSV with continuing index numbers. Returns count added."""
    existing = load_library(path)
    next_index = max((r.index or 0 for r in existing), default=0) + 1
    with Path(path).open("a", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        for i, row in enumerate(rows):
            row.index = next_index + i
            writer.writerow(row.to_csv_row())
    return len(rows)


def append_to_xlsx(rows: list[ScenarioRow], path: str | Path = LIBRARY_XLSX) -> int:
    """Append rows to the workbook's Scenario Library sheet (best-effort).

    Returns the number of rows added; 0 if ``openpyxl`` is unavailable or the
    file is missing.
    """
    try:
        import openpyxl
    except ImportError:
        return 0
    p = Path(path)
    if not p.exists():
        return 0
    wb = openpyxl.load_workbook(p)
    if LIBRARY_SHEET not in wb.sheetnames:
        return 0
    ws = wb[LIBRARY_SHEET]
    # next index = current max numeric value in column A
    max_idx = 0
    for (val,) in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if isinstance(val, int):
            max_idx = max(max_idx, val)
        elif isinstance(val, str) and val.strip().isdigit():
            max_idx = max(max_idx, int(val))
    for i, row in enumerate(rows):
        row.index = max_idx + 1 + i
        ws.append(row.to_csv_row())
    # best-effort: bump the Summary total
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        for r in summary.iter_rows(min_row=1, max_col=2):
            label, cell = r[0].value, r[1]
            if isinstance(label, str) and label.strip().lower() == "total scenarios":
                with contextlib.suppress(TypeError, ValueError):
                    cell.value = int(cell.value) + len(rows)
    wb.save(p)
    return len(rows)


def register_missing(
    *,
    csv_path: str | Path = LIBRARY_CSV,
    xlsx_path: str | Path = LIBRARY_XLSX,
    service_dir: str | Path = SERVICE_DIR,
) -> tuple[list[ScenarioRow], int]:
    """Add any repo scenarios missing from the library to the CSV and the xlsx.

    Returns ``(added_rows, xlsx_added)`` where ``xlsx_added`` is 0 when the
    workbook could not be updated.
    """
    library = load_library(csv_path)
    missing = missing_from_library(library, repo_scenarios(service_dir))
    if not missing:
        return [], 0
    append_to_csv(missing, csv_path)
    xlsx_added = append_to_xlsx([r.model_copy() for r in missing], xlsx_path)
    return missing, xlsx_added


__all__ = [
    "LIBRARY_CSV",
    "LIBRARY_HEADER",
    "LIBRARY_XLSX",
    "ScenarioRow",
    "append_to_csv",
    "append_to_xlsx",
    "find",
    "load_library",
    "missing_from_library",
    "register_missing",
    "repo_scenarios",
    "search",
]
